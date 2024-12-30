import os
import sys
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from tqdm import tqdm

def register_font():
    font_paths = [
        '/System/Library/Fonts/ヒラギノ角ゴシック W3.otf',
        '/System/Library/Fonts/AquaKana.ttc',
        '/System/Library/Fonts/AppleGothic.ttf',
        '/System/Library/Fonts/AppleSDGothicNeo.ttc',
        '/System/Library/Fonts/ヒラギノ明朝 ProN.ttc',
    ]
    
    for font_path in font_paths:
        try:
            pdfmetrics.registerFont(TTFont('JapaneseFont', font_path))
            print(f"Successfully registered font: {font_path}")
            return
        except Exception as e:
            print(f"Failed to register font {font_path}: {e}")
    
    print("Warning: No suitable Japanese font found. Using default font.")

def excel_to_pdf(excel_file):
    if not os.path.exists(excel_file):
        raise FileNotFoundError(f"指定されたExcelファイル '{excel_file}' が見つかりません。")

    register_font()

    wb = load_workbook(excel_file, data_only=True)
    book_name = os.path.splitext(os.path.basename(excel_file))[0]
    
    for sheet_name in tqdm(wb.sheetnames, desc="シートの処理"):
        sheet = wb[sheet_name]
        pdf_file = f"{book_name}_{sheet_name}.pdf"
        
        doc = SimpleDocTemplate(pdf_file, pagesize=landscape(letter))
        elements = []
        styles = getSampleStyleSheet()
        styles['Normal'].fontName = 'JapaneseFont'
        styles['Heading1'].fontName = 'JapaneseFont'

        elements.append(Paragraph(f"シート: {sheet_name}", styles['Heading1']))
        
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append([str(cell) if cell is not None else '' for cell in row])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'JapaneseFont'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
    
        doc.build(elements)
        print(f"PDFファイルを作成しました: {pdf_file}")

def main():
    if len(sys.argv) != 2:
        print("使用方法: python script.py <入力Excelファイルパス>")
        sys.exit(1)

    excel_file = sys.argv[1]

    try:
        excel_to_pdf(excel_file)
        print("全てのシートのPDF変換が完了しました。")
    except Exception as e:
        print(f"エラーが発生しました: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
